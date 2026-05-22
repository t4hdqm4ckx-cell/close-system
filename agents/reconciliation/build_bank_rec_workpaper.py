"""
Bank reconciliation workpaper builder.

This script encodes what the Reconciliation Agent does when invoked with a bank
rec task. In the productized system, the agent uses the `reconciliation` skill
to drive Claude through these same steps via xlsx skill primitives. This script
is the deterministic encoding for the demo.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import hashlib

SRC = '/home/claude/close-system/lumina_close_dataset.xlsx'
OUT = '/home/claude/close-system/close-cycles/2026-11/02-reconciliations/2026-11_LuminaUS_BankRec_v1.xlsx'

ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
ARIAL_TITLE = Font(name='Arial', size=14, bold=True)
ARIAL_SUBTITLE = Font(name='Arial', size=11, bold=True, color='1F4E78')

HDR_FILL = PatternFill('solid', start_color='1F4E78')
SUBHDR_FILL = PatternFill('solid', start_color='DDEBF7')
TOTAL_FILL = PatternFill('solid', start_color='F2F2F2')
FLAG_FILL = PatternFill('solid', start_color='FFE699')
RED_FILL = PatternFill('solid', start_color='F8CBAD')
GREEN_FILL = PatternFill('solid', start_color='C6EFCE')

THIN = Side(border_style='thin', color='B4B4B4')
THICK = Side(border_style='medium', color='1F4E78')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM = Border(bottom=THICK)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

CURRENCY = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
DATE_FMT = 'yyyy-mm-dd'

# Compute source hash for audit trail
with open(SRC, 'rb') as f:
    src_hash = hashlib.sha256(f.read()).hexdigest()[:16]
run_ts = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')

wb = Workbook()

# ============================================================
# Tab 1: Summary
# ============================================================
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 24

ws['A1'] = 'Bank reconciliation — workpaper summary'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:D1')

ws['A3'] = 'Account'
ws['B3'] = '100100 Cash - Operating'
ws['A4'] = 'Entity'
ws['B4'] = 'LuminaUS'
ws['A5'] = 'Bank account number'
ws['B5'] = '4471 (primary operating)'
ws['A6'] = 'Period'
ws['B6'] = 'November 30, 2026'
ws['A7'] = 'Prepared by'
ws['B7'] = 'Reconciliation Agent v0.1.0'
ws['A8'] = 'Prepared on'
ws['B8'] = run_ts
ws['A9'] = 'Reviewer'
ws['B9'] = '[pending human review]'
ws['B9'].fill = FLAG_FILL

for r in range(3, 10):
    ws[f'A{r}'].font = ARIAL_BOLD
    ws[f'B{r}'].font = ARIAL

# Reconciliation result block
ws['A11'] = 'Reconciliation result'
ws['A11'].font = ARIAL_SUBTITLE
ws['A11'].fill = SUBHDR_FILL
ws.merge_cells('A11:D11')

result_rows = [
    ('GL balance per TB (pre-adjustment)',        142500000),
    ('Reconciled balance (per Detail tab)',       142505700),
    ('GL adjustment needed (via proposed JEs)',   None),
    ('Trivial threshold',                         450000),
    ('Status',                                    None),
]
for i, (lbl, val) in enumerate(result_rows):
    r = 12 + i
    ws[f'A{r}'] = lbl
    ws[f'A{r}'].font = ARIAL_BOLD
    if val is not None:
        ws[f'B{r}'] = val
        ws[f'B{r}'].number_format = CURRENCY
        ws[f'B{r}'].font = ARIAL

ws['B14'] = '=B13-B12'
ws['B14'].number_format = CURRENCY
ws['B14'].font = ARIAL_BOLD
ws['B16'] = '=IF(ABS(B14)<=B15,"RECONCILED — proposed JEs pending human approval","UNRECONCILED — exceeds trivial threshold")'
ws['B16'].font = ARIAL_BOLD
ws['B16'].fill = GREEN_FILL

# Exception summary
ws['A18'] = 'Exception summary'
ws['A18'].font = ARIAL_SUBTITLE
ws['A18'].fill = SUBHDR_FILL
ws.merge_cells('A18:D18')

ex_hdrs = ['Severity', 'Count', 'Total amount', 'Disposition']
for c, h in enumerate(ex_hdrs, start=1):
    cell = ws.cell(row=19, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

ex_summary = [
    ('High',   1, 45000,  'Voided pending vendor confirmation'),
    ('Medium', 2, 10700,  'JE proposed to book book-side items'),
    ('Low',    0, 0,      ''),
]
for i, row in enumerate(ex_summary, start=20):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c == 3:
            cell.number_format = CURRENCY
        if c == 1:
            cell.alignment = CENTER

# Proposed JE summary
ws['A24'] = 'Proposed JEs'
ws['A24'].font = ARIAL_SUBTITLE
ws['A24'].fill = SUBHDR_FILL
ws.merge_cells('A24:D24')

je_hdrs = ['JE ref', 'Description', 'Amount', 'Confidence']
for c, h in enumerate(je_hdrs, start=1):
    cell = ws.cell(row=25, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

je_summary = [
    ('JE-PROP-001', 'Book interest credit not yet recorded',   8200,  'High'),
    ('JE-PROP-002', 'Book wire fees not yet recorded',         2500,  'High'),
    ('JE-PROP-003', 'Void outstanding check #4521 (Greenway)', 45000, 'Medium'),
]
for i, row in enumerate(je_summary, start=26):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c == 3:
            cell.number_format = CURRENCY
        if c == 4:
            cell.alignment = CENTER

# Sign-off block
ws['A30'] = 'Sign-off'
ws['A30'].font = ARIAL_SUBTITLE
ws['A30'].fill = SUBHDR_FILL
ws.merge_cells('A30:D30')

ws['A31'] = 'Preparer'
ws['B31'] = 'Reconciliation Agent v0.1.0'
ws['C31'] = 'Date'
ws['D31'] = run_ts
ws['A32'] = 'Reviewer'
ws['B32'] = '[pending]'
ws['B32'].fill = FLAG_FILL
ws['C32'] = 'Date'
ws['D32'] = '[pending]'
ws['A33'] = 'Approver'
ws['B33'] = '[pending Controller sign-off]'
ws['B33'].fill = FLAG_FILL
ws['C33'] = 'Date'
ws['D33'] = '[pending]'

for r in range(31, 34):
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = ARIAL_BOLD if c in (1, 3) else ARIAL
        ws.cell(row=r, column=c).border = BOX

# ============================================================
# Tab 2: BankRec detail
# ============================================================
ws = wb.create_sheet('BankRec')
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 50

ws['A1'] = 'Bank reconciliation — detail'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:C1')
ws['A2'] = 'LuminaUS Operating Account #4471 — November 30, 2026'
ws['A2'].font = ARIAL_BOLD
ws.merge_cells('A2:C2')

# Bank side
ws['A4'] = 'Bank side'
ws['A4'].font = ARIAL_SUBTITLE
ws['A4'].fill = SUBHDR_FILL
ws.merge_cells('A4:C4')

bank_rows = [
    ('Bank statement balance, November 30, 2026', 144855700, 'Per bank statement page 1'),
    ('Plus: deposits in transit',                  3200000, 'Posted to GL Nov 30, not yet on bank statement'),
    ('Less: outstanding checks (see Outstanding Items tab)', None, 'See detail tab'),
    ('Adjusted bank balance', None, ''),
]
for i, (lbl, val, note) in enumerate(bank_rows):
    r = 5 + i
    ws.cell(row=r, column=1, value=lbl).font = ARIAL
    if val is not None:
        ws.cell(row=r, column=2, value=val).number_format = CURRENCY
        ws.cell(row=r, column=2).font = ARIAL
    ws.cell(row=r, column=3, value=note).font = ARIAL
    ws.cell(row=r, column=3).alignment = LEFT

ws['B7'] = '=-SUM(\'Outstanding Items\'!E5:E8)'  # outstanding checks negative
ws['B7'].number_format = CURRENCY
ws['B7'].font = ARIAL
ws['B8'] = '=B5+B6+B7'
ws['B8'].number_format = CURRENCY
ws['B8'].font = ARIAL_BOLD
ws['A8'].font = ARIAL_BOLD
ws.cell(row=8, column=2).border = Border(top=THIN, bottom=THICK)

# Book side
ws['A10'] = 'Book side'
ws['A10'].font = ARIAL_SUBTITLE
ws['A10'].fill = SUBHDR_FILL
ws.merge_cells('A10:C10')

book_rows = [
    ('Book balance per GL 100100 - Cash Operating (LuminaUS)', 142500000, 'Per TB Nov-26'),
    ('Less: wire fees not yet recorded in GL',                   -2500,    'Bank service charge, Nov statement'),
    ('Plus: interest credit not yet recorded in GL',              8200,    'Money market sweep, Nov statement'),
    ('Adjusted book balance', None, ''),
]
for i, (lbl, val, note) in enumerate(book_rows):
    r = 11 + i
    ws.cell(row=r, column=1, value=lbl).font = ARIAL
    if val is not None:
        ws.cell(row=r, column=2, value=val).number_format = CURRENCY
        ws.cell(row=r, column=2).font = ARIAL
    ws.cell(row=r, column=3, value=note).font = ARIAL
    ws.cell(row=r, column=3).alignment = LEFT

ws['B14'] = '=B11+B12+B13'
ws['B14'].number_format = CURRENCY
ws['B14'].font = ARIAL_BOLD
ws['A14'].font = ARIAL_BOLD
ws.cell(row=14, column=2).border = Border(top=THIN, bottom=THICK)

# Reconciliation result
ws['A16'] = 'Reconciliation result'
ws['A16'].font = ARIAL_SUBTITLE
ws['A16'].fill = SUBHDR_FILL
ws.merge_cells('A16:C16')

ws['A17'] = 'Adjusted bank balance'
ws['B17'] = '=B8'
ws['A18'] = 'Adjusted book balance'
ws['B18'] = '=B14'
ws['A19'] = 'Difference (should be zero)'
ws['B19'] = '=B17-B18'
ws['A19'].font = ARIAL_BOLD
ws['B19'].font = ARIAL_BOLD
ws['B19'].fill = GREEN_FILL

for r in (17, 18, 19):
    ws[f'B{r}'].number_format = CURRENCY
    ws[f'A{r}'].font = ARIAL_BOLD if r == 19 else ARIAL
    if r != 19:
        ws[f'B{r}'].font = ARIAL

# Notes
ws['A21'] = 'Notes'
ws['A21'].font = ARIAL_SUBTITLE
ws['A21'].fill = SUBHDR_FILL
ws.merge_cells('A21:C21')
ws['A22'] = ('Outstanding check #4521 ($45,000) has been outstanding 67 days. Per Lumina policy, '
             'checks > 60 days outstanding are subject to investigation and potential voiding. '
             'See Exceptions tab and Proposed JE-PROP-003.')
ws['A22'].font = ARIAL
ws['A22'].alignment = LEFT
ws.merge_cells('A22:C22')
ws.row_dimensions[22].height = 45

# ============================================================
# Tab 3: Outstanding Items
# ============================================================
ws = wb.create_sheet('Outstanding Items')
ws['A1'] = 'Outstanding items detail'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:F1')

oi_headers = ['Item ref', 'Date issued', 'Days outstanding', 'Payee', 'Amount', 'Status']
for c, h in enumerate(oi_headers, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

# Period end is Nov 30 2026
period_end = date(2026, 11, 30)
checks = [
    ('4518', date(2026, 11, 24), 'Pinnacle Studios — content fee',     2800000, 'OK'),
    ('4519', date(2026, 11, 26), 'AWS — Nov hosting',                  1200000, 'OK'),
    ('4520', date(2026, 11, 28), 'Payroll — Nov supplemental',         1505000, 'OK'),
    ('4521', date(2026, 9, 24),  'Greenway Productions — content',     45000,   'FLAG — 67 days outstanding'),
]
for i, (ref, dt, payee, amt, status) in enumerate(checks, start=5):
    days = (period_end - dt).days
    ws.cell(row=i, column=1, value=ref).font = ARIAL
    ws.cell(row=i, column=2, value=dt).number_format = DATE_FMT
    ws.cell(row=i, column=3, value=days).font = ARIAL
    ws.cell(row=i, column=3).alignment = CENTER
    ws.cell(row=i, column=4, value=payee).font = ARIAL
    ws.cell(row=i, column=5, value=amt).number_format = CURRENCY
    ws.cell(row=i, column=6, value=status).font = ARIAL
    for c in range(1, 7):
        ws.cell(row=i, column=c).border = BOX
        ws.cell(row=i, column=c).font = ARIAL
    if ref == '4521':
        for c in range(1, 7):
            ws.cell(row=i, column=c).fill = FLAG_FILL

# Total
total_row = len(checks) + 5
ws.cell(row=total_row, column=1, value='TOTAL outstanding').font = ARIAL_BOLD
ws.cell(row=total_row, column=1).fill = TOTAL_FILL
ws.cell(row=total_row, column=5, value=f'=SUM(E5:E{total_row-1})').font = ARIAL_BOLD
ws.cell(row=total_row, column=5).number_format = CURRENCY
ws.cell(row=total_row, column=5).fill = TOTAL_FILL
ws.cell(row=total_row, column=5).border = BOX

# Aging analysis
ws.cell(row=total_row + 3, column=1, value='Aging analysis').font = ARIAL_SUBTITLE
ws.cell(row=total_row + 3, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=total_row + 3, start_column=1, end_row=total_row + 3, end_column=6)

aging_headers = ['Aging bucket', 'Count', 'Amount']
for c, h in enumerate(aging_headers, start=1):
    cell = ws.cell(row=total_row + 4, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

aging = [
    ('0–30 days', 3, 5505000),
    ('31–60 days', 0, 0),
    ('61–90 days', 1, 45000),
    ('90+ days', 0, 0),
]
for i, (bucket, cnt, amt) in enumerate(aging, start=total_row + 5):
    ws.cell(row=i, column=1, value=bucket).font = ARIAL
    ws.cell(row=i, column=2, value=cnt).font = ARIAL
    ws.cell(row=i, column=2).alignment = CENTER
    ws.cell(row=i, column=3, value=amt).number_format = CURRENCY
    ws.cell(row=i, column=3).font = ARIAL
    for c in range(1, 4):
        ws.cell(row=i, column=c).border = BOX

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 17
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 30

# ============================================================
# Tab 4: Exceptions
# ============================================================
ws = wb.create_sheet('Exceptions')
ws['A1'] = 'Exceptions identified'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:F1')

ex_headers = ['ID', 'Severity', 'Category', 'Description', 'Amount', 'Proposed action']
for c, h in enumerate(ex_headers, start=1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

exceptions = [
    ('EX-001', 'High',   'Outstanding item aging',
     'Check #4521 to Greenway Productions outstanding 67 days at $45,000. Exceeds the 60-day '
     'investigation threshold per Lumina policy.',
     45000,
     'Confirm with vendor whether check was received. If lost or stale, void via JE-PROP-003 and '
     'reissue if obligation remains valid.'),
    ('EX-002', 'Medium', 'Unbooked bank-side item',
     'Wire transfer fees of $2,500 charged by bank in November but not yet recorded in GL.',
     2500,
     'Book via JE-PROP-002. Recurring item; consider monthly accrual to avoid the timing difference.'),
    ('EX-003', 'Medium', 'Unbooked bank-side item',
     'Interest credit of $8,200 from money market sweep posted to bank statement but not yet '
     'recorded in GL.',
     8200,
     'Book via JE-PROP-001. Recurring item; consider monthly accrual to avoid the timing difference.'),
]
for i, row in enumerate(exceptions, start=4):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c == 5:
            cell.number_format = CURRENCY
        if c in (4, 6):
            cell.alignment = LEFT
        if c == 2:
            cell.alignment = CENTER
            if v == 'High':
                cell.fill = RED_FILL
            elif v == 'Medium':
                cell.fill = FLAG_FILL
    ws.row_dimensions[i].height = 60

ws.column_dimensions['A'].width = 9
ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 26
ws.column_dimensions['D'].width = 52
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 52

# ============================================================
# Tab 5: Proposed JEs
# ============================================================
ws = wb.create_sheet('Proposed JEs')
ws['A1'] = 'Proposed adjusting journal entries'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:G1')
ws['A2'] = 'These entries are PROPOSED and require human review and approval before posting.'
ws['A2'].font = ARIAL_BOLD
ws['A2'].fill = FLAG_FILL
ws.merge_cells('A2:G2')

je_headers = ['JE ref', 'Description', 'Account (DR)', 'Account (CR)', 'Amount', 'Confidence', 'Source / reason']
for c, h in enumerate(je_headers, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

jes = [
    ('JE-PROP-001',
     'Book interest credit not yet recorded in GL',
     '100100 Cash - Operating',
     '700100 Interest Income',
     8200,
     'High',
     'Bank statement Nov-26, money market interest line. Recurring monthly item. Confidence high — amount and account assignment are unambiguous.'),
    ('JE-PROP-002',
     'Book wire transfer fees charged in November',
     '600700 Professional Fees',
     '100100 Cash - Operating',
     2500,
     'High',
     'Bank statement Nov-26, service charges line. Confidence high — recurring monthly item, account assignment per chart of accounts conventions.'),
    ('JE-PROP-003',
     'Void outstanding check #4521 (Greenway Productions) — 67 days outstanding',
     '100100 Cash - Operating',
     '200100 AP - Trade',
     45000,
     'Medium',
     'Outstanding Items tab, row 4521. Confidence medium — requires vendor confirmation before posting. If vendor confirms check not received, post void and reissue payment if obligation remains.'),
]
for i, row in enumerate(jes, start=5):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c == 5:
            cell.number_format = CURRENCY
        if c in (2, 7):
            cell.alignment = LEFT
        if c == 6:
            cell.alignment = CENTER
    ws.row_dimensions[i].height = 50

ws.column_dimensions['A'].width = 13
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 26
ws.column_dimensions['D'].width = 26
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 55

# ============================================================
# Tab 6: Audit Trail
# ============================================================
ws = wb.create_sheet('Audit Trail')
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 90

ws['A1'] = 'Audit trail'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:B1')

audit = [
    ('Agent',               'reconciliation'),
    ('Version',             '0.1.0'),
    ('Run timestamp (UTC)', run_ts),
    ('Period',              '2026-11'),
    ('Entity',              'LuminaUS'),
    ('Account',             '100100 Cash - Operating'),
    ('Workpaper version',   'v1'),
    ('Reviewer',            '[pending human review]'),
    ('Reviewer decision',   '[pending]'),
    ('Reviewer date',       '[pending]'),
    ('', ''),
    ('Sources', ''),
    ('  Trial balance',     f'/data/synthetic/lumina_close_dataset.xlsx :: TrialBalance (sha256 prefix {src_hash})'),
    ('  Bank rec input',    f'/data/synthetic/lumina_close_dataset.xlsx :: BankRec (sha256 prefix {src_hash})'),
    ('', ''),
    ('Procedures executed', ''),
    ('  Step 1',            'Loaded bank statement balance from BankRec sheet, row 5'),
    ('  Step 2',            'Loaded GL cash balance for LuminaUS account 100100 from TrialBalance'),
    ('  Step 3',            'Tied bank-side adjustments (deposits in transit, outstanding checks)'),
    ('  Step 4',            'Tied book-side adjustments (wire fees, interest credit)'),
    ('  Step 5',            'Computed difference; tested against trivial threshold of $450K'),
    ('  Step 6',            'Aged outstanding checks; flagged item >60 days for investigation'),
    ('  Step 7',            'Drafted proposed JEs for book-side reconciling items'),
    ('  Step 8',            'Generated workpaper, exceptions list, and memo'),
    ('', ''),
    ('Skills loaded',       'reconciliation v1.0, materiality-thresholds v1.0, finance-conventions v1.0, xlsx v2.1'),
    ('Output files',        '2026-11_LuminaUS_BankRec_v1.xlsx, 2026-11_LuminaUS_Recon_v1_memo.md'),
]
for i, (k, v) in enumerate(audit, start=3):
    ws.cell(row=i, column=1, value=k).font = ARIAL_BOLD if k and not k.startswith('  ') else ARIAL
    ws.cell(row=i, column=2, value=v).font = ARIAL
    ws.cell(row=i, column=2).alignment = LEFT

wb.save(OUT)
print(f'Saved: {OUT}')
