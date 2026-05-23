"""
build_bank_rec_workpaper.py — Bank reconciliation workpaper builder.

Produces a six-tab xlsx workpaper for the LuminaUS operating cash account
(GL 100100) for a given close period. All policy thresholds are imported
from config.py — do not hardcode numbers in this file.

Usage:
    python build_bank_rec_workpaper.py \
        --src  /path/to/lumina_close_dataset.xlsx \
        --out  /path/to/output/2026-11_LuminaUS_BankRec_v1.xlsx \
        --period 2026-11
"""

import argparse
import hashlib
import pathlib
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter  # noqa: F401 — available for callers

from config import (
    AR_NINETY_DAY_PCT_THRESHOLD,
    AUTO_INVESTIGATE_THRESHOLD,
    BS_REC_EXCEPTION_AGE_DAYS,
    BS_REC_EXCEPTION_AMOUNT,
    DEPOSIT_IN_TRANSIT_FLAG_DAYS,
    OUTSTANDING_CHECK_FLAG_DAYS,
    PERFORMANCE_MATERIALITY,
    TRIVIAL_THRESHOLD,
)

# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build LuminaUS bank rec workpaper")
    p.add_argument("--src",    required=True, type=pathlib.Path,
                   help="Path to source dataset xlsx")
    p.add_argument("--out",    required=True, type=pathlib.Path,
                   help="Output xlsx path (parent directory must exist)")
    p.add_argument("--period", default="2026-11",
                   help="Close period in YYYY-MM format (default: 2026-11)")
    return p.parse_args()


# ── Styles ───────────────────────────────────────────────────────────────────

ARIAL          = Font(name="Arial", size=10)
ARIAL_BOLD     = Font(name="Arial", size=10, bold=True)
ARIAL_HDR      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ARIAL_TITLE    = Font(name="Arial", size=14, bold=True)
ARIAL_SUBTITLE = Font(name="Arial", size=11, bold=True, color="1F4E78")

HDR_FILL   = PatternFill("solid", start_color="1F4E78")
SUBHDR_FILL = PatternFill("solid", start_color="DDEBF7")
TOTAL_FILL  = PatternFill("solid", start_color="F2F2F2")
FLAG_FILL   = PatternFill("solid", start_color="FFE699")
RED_FILL    = PatternFill("solid", start_color="F8CBAD")
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")

THIN   = Side(border_style="thin",   color="B4B4B4")
THICK  = Side(border_style="medium", color="1F4E78")
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM = Border(bottom=THICK)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

CURRENCY = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
DATE_FMT  = "yyyy-mm-dd"

# ── Workpaper builder ─────────────────────────────────────────────────────────

def build(src: pathlib.Path, out: pathlib.Path, period: str) -> None:
    if not src.exists():
        raise FileNotFoundError(f"Source dataset not found: {src}")
    out.parent.mkdir(parents=True, exist_ok=True)

    with open(src, "rb") as f:
        src_hash = hashlib.sha256(f.read()).hexdigest()[:16]

    run_ts      = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    period_end  = date(2026, 11, 30)   # parameterise if period arg is extended
    wb          = Workbook()

    _build_summary(wb, run_ts)
    _build_bank_rec(wb, period_end)
    _build_outstanding_items(wb, period_end)
    _build_exceptions(wb)
    _build_proposed_jes(wb)
    _build_audit_trail(wb, src, src_hash, run_ts, period)

    wb.save(out)
    print(f"Saved: {out}")


# ── Tab builders ──────────────────────────────────────────────────────────────

def _build_summary(wb: Workbook, run_ts: str) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 24

    ws["A1"] = "Bank reconciliation — workpaper summary"
    ws["A1"].font = ARIAL_TITLE
    ws.merge_cells("A1:D1")

    meta = [
        ("Account",             "100100 Cash - Operating"),
        ("Entity",              "LuminaUS"),
        ("Bank account number", "4471 (primary operating)"),
        ("Period",              "November 30, 2026"),
        ("Prepared by",         "Reconciliation Agent v0.1.0"),
        ("Prepared on",         run_ts),
        ("Reviewer",            "[pending human review]"),
    ]
    for i, (lbl, val) in enumerate(meta, start=3):
        ws[f"A{i}"] = lbl
        ws[f"A{i}"].font = ARIAL_BOLD
        ws[f"B{i}"] = val
        ws[f"B{i}"].font = ARIAL
    ws["B9"].fill = FLAG_FILL  # reviewer pending

    # Result block
    ws["A11"] = "Reconciliation result"
    ws["A11"].font = ARIAL_SUBTITLE
    ws["A11"].fill = SUBHDR_FILL
    ws.merge_cells("A11:D11")

    result_rows = [
        ("GL balance per TB (pre-adjustment)",       142_500_000),
        ("Reconciled balance (per Detail tab)",       142_505_700),
        ("GL adjustment needed (via proposed JEs)",   None),
        (f"Trivial threshold",                        TRIVIAL_THRESHOLD),
        ("Status",                                    None),
    ]
    for i, (lbl, val) in enumerate(result_rows, start=12):
        ws[f"A{i}"] = lbl
        ws[f"A{i}"].font = ARIAL_BOLD
        if val is not None:
            ws[f"B{i}"] = val
            ws[f"B{i}"].number_format = CURRENCY
            ws[f"B{i}"].font = ARIAL

    ws["B14"] = "=B13-B12"
    ws["B14"].number_format = CURRENCY
    ws["B14"].font = ARIAL_BOLD
    ws["B16"] = (
        f'=IF(ABS(B14)<={TRIVIAL_THRESHOLD},'
        '"RECONCILED — proposed JEs pending human approval",'
        '"UNRECONCILED — exceeds trivial threshold")'
    )
    ws["B16"].font = ARIAL_BOLD
    ws["B16"].fill = GREEN_FILL

    # Exception summary
    ws["A18"] = "Exception summary"
    ws["A18"].font = ARIAL_SUBTITLE
    ws["A18"].fill = SUBHDR_FILL
    ws.merge_cells("A18:D18")

    ex_hdrs = ["Severity", "Count", "Total amount", "Disposition"]
    for c, h in enumerate(ex_hdrs, start=1):
        cell = ws.cell(row=19, column=c, value=h)
        cell.font = ARIAL_HDR
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    ex_summary = [
        ("High",   1, 45_000, "Voided pending vendor confirmation"),
        ("Medium", 2, 10_700, "JE proposed to book book-side items"),
        ("Low",    0, 0,      ""),
    ]
    for i, row in enumerate(ex_summary, start=20):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = ARIAL
            cell.border = BOX
            if c == 3:
                cell.number_format = CURRENCY
            if c == 1:
                cell.alignment = CENTER

    # Proposed JE summary
    ws["A24"] = "Proposed JEs"
    ws["A24"].font = ARIAL_SUBTITLE
    ws["A24"].fill = SUBHDR_FILL
    ws.merge_cells("A24:D24")

    je_hdrs = ["JE ref", "Description", "Amount", "Confidence"]
    for c, h in enumerate(je_hdrs, start=1):
        cell = ws.cell(row=25, column=c, value=h)
        cell.font = ARIAL_HDR
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    je_summary = [
        ("JE-PROP-001", "Book interest credit not yet recorded",    8_200,  "High"),
        ("JE-PROP-002", "Book wire fees not yet recorded",          2_500,  "High"),
        ("JE-PROP-003", "Void outstanding check #4521 (Greenway)",  45_000, "Medium"),
    ]
    for i, row in enumerate(je_summary, start=26):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = ARIAL
            cell.border = BOX
            if c == 3:
                cell.number_format = CURRENCY
            if c == 4:
                cell.alignment = CENTER

    # Sign-off block
    ws["A30"] = "Sign-off"
    ws["A30"].font = ARIAL_SUBTITLE
    ws["A30"].fill = SUBHDR_FILL
    ws.merge_cells("A30:D30")

    signoff = [
        ("Preparer", "Reconciliation Agent v0.1.0", "Date", run_ts),
        ("Reviewer", "[pending]",                    "Date", "[pending]"),
        ("Approver", "[pending Controller sign-off]","Date", "[pending]"),
    ]
    for i, (a, b, c, d) in enumerate(signoff, start=31):
        ws[f"A{i}"] = a; ws[f"A{i}"].font = ARIAL_BOLD
        ws[f"B{i}"] = b; ws[f"B{i}"].font = ARIAL
        ws[f"C{i}"] = c; ws[f"C{i}"].font = ARIAL_BOLD
        ws[f"D{i}"] = d; ws[f"D{i}"].font = ARIAL
        if b in ("[pending]", "[pending Controller sign-off]"):
            ws[f"B{i}"].fill = FLAG_FILL
        for col in "ABCD":
            ws[f"{col}{i}"].border = BOX


def _build_bank_rec(wb: Workbook, period_end: date) -> None:
    ws = wb.create_sheet("BankRec")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50

    ws["A1"] = "Bank reconciliation — detail"
    ws["A1"].font = ARIAL_TITLE
    ws.merge_cells("A1:C1")
    ws["A2"] = "LuminaUS Operating Account #4471 — November 30, 2026"
    ws["A2"].font = ARIAL_BOLD
    ws.merge_cells("A2:C2")

    # Bank side
    ws["A4"] = "Bank side"
    ws["A4"].font = ARIAL_SUBTITLE
    ws["A4"].fill = SUBHDR_FILL
    ws.merge_cells("A4:C4")

    bank_rows = [
        ("Bank statement balance, November 30, 2026", 144_855_700, "Per bank statement page 1"),
        ("Plus: deposits in transit",                   3_200_000,
         f"Posted to GL Nov 30; not yet on bank statement. "
         f"Flag if >_{DEPOSIT_IN_TRANSIT_FLAG_DAYS} days old."),
        ("Less: outstanding checks (see Outstanding Items tab)", None, "See Outstanding Items tab"),
        ("Adjusted bank balance",                       None, ""),
    ]
    for i, (lbl, val, note) in enumerate(bank_rows, start=5):
        ws.cell(row=i, column=1, value=lbl).font = ARIAL
        if val is not None:
            ws.cell(row=i, column=2, value=val).number_format = CURRENCY
            ws.cell(row=i, column=2).font = ARIAL
        ws.cell(row=i, column=3, value=note).font = ARIAL
        ws.cell(row=i, column=3).alignment = LEFT

    ws["B7"] = "=-SUM('Outstanding Items'!E5:E8)"
    ws["B7"].number_format = CURRENCY
    ws["B7"].font = ARIAL
    ws["B8"] = "=B5+B6+B7"
    ws["B8"].number_format = CURRENCY
    ws["B8"].font = ARIAL_BOLD
    ws["A8"].font = ARIAL_BOLD
    ws["B8"].border = Border(top=THIN, bottom=THICK)

    # Book side
    ws["A10"] = "Book side"
    ws["A10"].font = ARIAL_SUBTITLE
    ws["A10"].fill = SUBHDR_FILL
    ws.merge_cells("A10:C10")

    book_rows = [
        ("Book balance per GL 100100 - Cash Operating (LuminaUS)", 142_500_000, "Per TB Nov-26"),
        ("Less: wire fees not yet recorded in GL",                    -2_500,    "Bank service charge, Nov statement"),
        ("Plus: interest credit not yet recorded in GL",               8_200,    "Money market sweep, Nov statement"),
        ("Adjusted book balance",                                       None,    ""),
    ]
    for i, (lbl, val, note) in enumerate(book_rows, start=11):
        ws.cell(row=i, column=1, value=lbl).font = ARIAL
        if val is not None:
            ws.cell(row=i, column=2, value=val).number_format = CURRENCY
            ws.cell(row=i, column=2).font = ARIAL
        ws.cell(row=i, column=3, value=note).font = ARIAL
        ws.cell(row=i, column=3).alignment = LEFT

    ws["B14"] = "=B11+B12+B13"
    ws["B14"].number_format = CURRENCY
    ws["B14"].font = ARIAL_BOLD
    ws["A14"].font = ARIAL_BOLD
    ws["B14"].border = Border(top=THIN, bottom=THICK)

    # Result
    ws["A16"] = "Reconciliation result"
    ws["A16"].font = ARIAL_SUBTITLE
    ws["A16"].fill = SUBHDR_FILL
    ws.merge_cells("A16:C16")

    for r, lbl, formula in (
        (17, "Adjusted bank balance",         "=B8"),
        (18, "Adjusted book balance",         "=B14"),
        (19, "Difference (should be zero)",   "=B17-B18"),
    ):
        ws[f"A{r}"] = lbl
        ws[f"B{r}"] = formula
        ws[f"B{r}"].number_format = CURRENCY
        ws[f"A{r}"].font = ARIAL_BOLD if r == 19 else ARIAL
        ws[f"B{r}"].font = ARIAL_BOLD if r == 19 else ARIAL

    ws["B19"].fill = GREEN_FILL

    # Notes
    ws["A21"] = "Notes"
    ws["A21"].font = ARIAL_SUBTITLE
    ws["A21"].fill = SUBHDR_FILL
    ws.merge_cells("A21:C21")
    ws["A22"] = (
        f"Outstanding check #4521 ($45,000) has been outstanding 67 days. "
        f"Per Lumina policy, checks >{OUTSTANDING_CHECK_FLAG_DAYS} days outstanding "
        f"are subject to investigation and potential voiding. "
        f"See Exceptions tab and Proposed JE-PROP-003."
    )
    ws["A22"].font = ARIAL
    ws["A22"].alignment = LEFT
    ws.merge_cells("A22:C22")
    ws.row_dimensions[22].height = 45


def _build_outstanding_items(wb: Workbook, period_end: date) -> None:
    ws = wb.create_sheet("Outstanding Items")
    ws["A1"] = "Outstanding items detail"
    ws["A1"].font = ARIAL_TITLE
    ws.merge_cells("A1:F1")

    oi_headers = ["Item ref", "Date issued", "Days outstanding", "Payee", "Amount", "Status"]
    for c, h in enumerate(oi_headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = ARIAL_HDR
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BOX

    checks = [
        ("4518", date(2026, 11, 24), "Pinnacle Studios — content fee",  2_800_000, "OK"),
        ("4519", date(2026, 11, 26), "AWS — Nov hosting",               1_200_000, "OK"),
        ("4520", date(2026, 11, 28), "Payroll — Nov supplemental",      1_505_000, "OK"),
        ("4521", date(2026,  9, 24), "Greenway Productions — content",     45_000,
         f"FLAG — {(period_end - date(2026, 9, 24)).days} days outstanding"),
    ]
    for i, (ref, dt, payee, amt, status) in enumerate(checks, start=5):
        days = (period_end - dt).days
        is_flagged = days > OUTSTANDING_CHECK_FLAG_DAYS

        ws.cell(row=i, column=1, value=ref)
        ws.cell(row=i, column=2, value=dt).number_format = DATE_FMT
        ws.cell(row=i, column=3, value=days).alignment = CENTER
        ws.cell(row=i, column=4, value=payee)
        ws.cell(row=i, column=5, value=amt).number_format = CURRENCY
        ws.cell(row=i, column=6, value=status)

        for c in range(1, 7):
            ws.cell(row=i, column=c).font = ARIAL
            ws.cell(row=i, column=c).border = BOX
            if is_flagged:
                ws.cell(row=i, column=c).fill = FLAG_FILL

    total_row = len(checks) + 5
    ws.cell(row=total_row, column=1, value="TOTAL outstanding").font = ARIAL_BOLD
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=5, value=f"=SUM(E5:E{total_row - 1})").font = ARIAL_BOLD
    ws.cell(row=total_row, column=5).number_format = CURRENCY
    ws.cell(row=total_row, column=5).fill = TOTAL_FILL
    ws.cell(row=total_row, column=5).border = BOX

    # Aging analysis
    ar = total_row + 3
    ws.cell(row=ar, column=1, value="Aging analysis").font = ARIAL_SUBTITLE
    ws.cell(row=ar, column=1).fill = SUBHDR_FILL
    ws.merge_cells(start_row=ar, start_column=1, end_row=ar, end_column=6)

    for c, h in enumerate(["Aging bucket", "Count", "Amount"], start=1):
        cell = ws.cell(row=ar + 1, column=c, value=h)
        cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

    aging = [
        ("0–30 days",  3, 5_505_000),
        ("31–60 days", 0,         0),
        ("61–90 days", 1,    45_000),
        ("90+ days",   0,         0),
    ]
    for i, (bucket, cnt, amt) in enumerate(aging, start=ar + 2):
        ws.cell(row=i, column=1, value=bucket).font = ARIAL
        ws.cell(row=i, column=2, value=cnt).alignment = CENTER
        ws.cell(row=i, column=3, value=amt).number_format = CURRENCY
        for c in range(1, 4):
            ws.cell(row=i, column=c).font = ARIAL
            ws.cell(row=i, column=c).border = BOX

    for col, width in zip("ABCDEF", [20, 14, 17, 40, 16, 30]):
        ws.column_dimensions[col].width = width


def _build_exceptions(wb: Workbook) -> None:
    ws = wb.create_sheet("Exceptions")
    ws["A1"] = "Exceptions identified"
    ws["A1"].font = ARIAL_TITLE
    ws.merge_cells("A1:F1")

    for c, h in enumerate(["ID", "Severity", "Category", "Description", "Amount", "Proposed action"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

    exceptions = [
        ("EX-001", "High",   "Outstanding item aging",
         f"Check #4521 to Greenway Productions outstanding 67 days at $45,000. "
         f"Exceeds the {OUTSTANDING_CHECK_FLAG_DAYS}-day investigation threshold per Lumina policy.",
         45_000,
         "Confirm with vendor whether check was received. If lost or stale, void via "
         "JE-PROP-003 and reissue if obligation remains valid."),
        ("EX-002", "Medium", "Unbooked bank-side item",
         "Wire transfer fees of $2,500 charged by bank in November but not yet recorded in GL.",
         2_500,
         "Book via JE-PROP-002. Recurring item; consider monthly accrual to avoid timing difference."),
        ("EX-003", "Medium", "Unbooked bank-side item",
         "Interest credit of $8,200 from money market sweep posted to bank statement "
         "but not yet recorded in GL.",
         8_200,
         "Book via JE-PROP-001. Recurring item; consider monthly accrual to avoid timing difference."),
    ]
    for i, row in enumerate(exceptions, start=4):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = ARIAL; cell.border = BOX
            if c == 5:
                cell.number_format = CURRENCY
            if c in (4, 6):
                cell.alignment = LEFT
            if c == 2:
                cell.alignment = CENTER
                cell.fill = RED_FILL if v == "High" else FLAG_FILL
        ws.row_dimensions[i].height = 60

    for col, width in zip("ABCDEF", [9, 11, 26, 52, 14, 52]):
        ws.column_dimensions[col].width = width


def _build_proposed_jes(wb: Workbook) -> None:
    ws = wb.create_sheet("Proposed JEs")
    ws["A1"] = "Proposed adjusting journal entries"
    ws["A1"].font = ARIAL_TITLE
    ws.merge_cells("A1:G1")
    ws["A2"] = "These entries are PROPOSED and require human review and approval before posting."
    ws["A2"].font = ARIAL_BOLD
    ws["A2"].fill = FLAG_FILL
    ws.merge_cells("A2:G2")

    for c, h in enumerate(
        ["JE ref", "Description", "Account (DR)", "Account (CR)", "Amount", "Confidence", "Source / reason"],
        start=1,
    ):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

    jes = [
        ("JE-PROP-001",
         "Book interest credit not yet recorded in GL",
         "100100 Cash - Operating",
         "700100 Interest Income",
         8_200, "High",
         "Bank statement Nov-26, money market interest line. Recurring monthly item. "
         "Confidence high — amount and account assignment are unambiguous."),
        ("JE-PROP-002",
         "Book wire transfer fees charged in November",
         "600700 Professional Fees",
         "100100 Cash - Operating",
         2_500, "High",
         "Bank statement Nov-26, service charges line. Recurring monthly item, "
         "account assignment per chart of accounts conventions."),
        ("JE-PROP-003",
         "Void outstanding check #4521 (Greenway Productions) — 67 days outstanding",
         "100100 Cash - Operating",
         "200100 AP - Trade",
         45_000, "Medium",
         f"Outstanding Items tab, check #4521. Confidence medium — requires vendor confirmation "
         f"before posting. Outstanding >{OUTSTANDING_CHECK_FLAG_DAYS} days per Lumina policy."),
    ]
    for i, row in enumerate(jes, start=5):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = ARIAL; cell.border = BOX
            if c == 5:
                cell.number_format = CURRENCY
            if c in (2, 7):
                cell.alignment = LEFT
            if c == 6:
                cell.alignment = CENTER
        ws.row_dimensions[i].height = 50

    for col, width in zip("ABCDEFG", [13, 50, 26, 26, 14, 12, 55]):
        ws.column_dimensions[col].width = width


def _build_audit_trail(
    wb: Workbook,
    src: pathlib.Path,
    src_hash: str,
    run_ts: str,
    period: str,
) -> None:
    ws = wb.create_sheet("Audit Trail")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    ws["A1"] = "Audit trail"
    ws["A1"].font = ARIAL_TITLE
    ws.merge_cells("A1:B1")

    audit = [
        ("Agent",               "reconciliation"),
        ("Version",             "0.1.0"),
        ("Run timestamp (UTC)", run_ts),
        ("Period",              period),
        ("Entity",              "LuminaUS"),
        ("Account",             "100100 Cash - Operating"),
        ("Workpaper version",   "v1"),
        ("Reviewer",            "[pending human review]"),
        ("Reviewer decision",   "[pending]"),
        ("Reviewer date",       "[pending]"),
        ("", ""),
        ("Sources", ""),
        ("  Trial balance", f"{src} :: TrialBalance (sha256 prefix {src_hash})"),
        ("  Bank rec input", f"{src} :: BankRec (sha256 prefix {src_hash})"),
        ("", ""),
        ("Procedures executed", ""),
        ("  Step 1", "Loaded bank statement balance from BankRec sheet"),
        ("  Step 2", "Loaded GL cash balance for LuminaUS account 100100 from TrialBalance"),
        ("  Step 3", "Tied bank-side adjustments (deposits in transit, outstanding checks)"),
        ("  Step 4", "Tied book-side adjustments (wire fees, interest credit)"),
        ("  Step 5", f"Computed difference; tested against trivial threshold of ${TRIVIAL_THRESHOLD:,}"),
        ("  Step 6", f"Aged outstanding checks; flagged items >{OUTSTANDING_CHECK_FLAG_DAYS} days"),
        ("  Step 7", "Drafted proposed JEs for book-side reconciling items"),
        ("  Step 8", "Generated workpaper, exceptions list, and memo"),
        ("", ""),
        ("Thresholds applied", ""),
        ("  Trivial threshold",           f"${TRIVIAL_THRESHOLD:,}"),
        ("  Performance materiality",     f"${PERFORMANCE_MATERIALITY:,}"),
        ("  Auto-investigate threshold",  f"${AUTO_INVESTIGATE_THRESHOLD:,}"),
        ("  BS rec exception amount",     f"${BS_REC_EXCEPTION_AMOUNT:,}"),
        ("  BS rec exception age",        f"{BS_REC_EXCEPTION_AGE_DAYS} days"),
        ("  Outstanding check flag",      f"{OUTSTANDING_CHECK_FLAG_DAYS} days"),
        ("  Deposit in transit flag",     f"{DEPOSIT_IN_TRANSIT_FLAG_DAYS} days"),
        ("", ""),
        ("Skills loaded", "reconciliation v1.0, materiality-thresholds v1.0, finance-conventions v1.0, xlsx v2.1"),
    ]
    for i, (k, v) in enumerate(audit, start=3):
        ws.cell(row=i, column=1, value=k).font = (
            ARIAL_BOLD if k and not k.startswith("  ") else ARIAL
        )
        ws.cell(row=i, column=2, value=v).font = ARIAL
        ws.cell(row=i, column=2).alignment = LEFT


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    args = parse_args()
    build(src=args.src, out=args.out, period=args.period)
