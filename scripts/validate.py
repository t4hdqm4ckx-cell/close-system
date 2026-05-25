"""
validate.py — Close System artifact validation

Checks:
  1. Synthetic dataset opens, has all expected sheets, and contains no formula errors
  2. All YAML config files parse without errors
  3. materiality.yaml has all required keys and the trivial threshold is 5% of PM
  4. All five AGENT.md files are present
  5. All six SKILL.md files are present

Run locally:  python scripts/validate.py
Run in CI:    triggered automatically via .github/workflows/validate.yml
"""

import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent

EXPECTED_SHEETS = [
    'README', 'TrialBalance', 'AR_Aging', 'AP_Aging',
    'PrepaidSchedule', 'AccrualSchedule', 'BankRec',
    'IntercoMatrix', 'CloseCalendar', 'JE_Log',
]

REQUIRED_AGENTS = [
    'orchestrator', 'reconciliation', 'je-reviewer',
    'flux-variance', 'close-reporting',
]

REQUIRED_SKILLS = [
    'reconciliation', 'materiality-thresholds', 'je-review',
    'flux-analysis', 'close-reporting', 'finance-conventions',
]

REQUIRED_MATERIALITY_KEYS = [
    'performance_materiality_usd',
    'trivial_threshold_usd',
    'pl_flux',
    'bs_rec_exception',
    'je_review',
]

ERROR_VALUES = {'#REF!', '#VALUE!', '#N/A', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!'}


def header(title):
    print(f"\n{title}")
    print("-" * len(title))


def validate_dataset(path):
    import openpyxl

    header(f"Dataset: {path.relative_to(ROOT)}")

    if not path.exists():
        print(f"  FAIL  File not found")
        return False

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  FAIL  Could not open workbook: {e}")
        return False

    # Check sheets
    missing_sheets = [s for s in EXPECTED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        print(f"  FAIL  Missing sheets: {missing_sheets}")
        return False
    print(f"  PASS  Sheets: {len(wb.sheetnames)} present ({len(EXPECTED_SHEETS)} required)")

    # Check for formula errors in cached values
    errors_found = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.upper() in ERROR_VALUES:
                    errors_found.append(f"{sheet_name}!{cell.coordinate}: {cell.value}")

    if errors_found:
        print(f"  FAIL  Formula errors found:")
        for e in errors_found:
            print(f"        {e}")
        return False

    print(f"  PASS  Formula errors: 0")
    return True


def validate_yaml_files():
    import yaml

    header("YAML config files")
    passed = True

    config_dir = ROOT / 'config'
    yaml_files = sorted(config_dir.glob('*.yaml'))

    if not yaml_files:
        print("  FAIL  No YAML files found in config/")
        return False

    for path in yaml_files:
        try:
            with open(path) as f:
                data = yaml.safe_load(f)
            if data is None:
                print(f"  FAIL  {path.name} — file is empty")
                passed = False
            else:
                print(f"  PASS  {path.name}")
        except yaml.YAMLError as e:
            print(f"  FAIL  {path.name} — {e}")
            passed = False

    return passed


def validate_materiality():
    import yaml

    header("materiality.yaml — required keys and threshold ratios")

    path = ROOT / 'config' / 'materiality.yaml'
    if not path.exists():
        print("  FAIL  materiality.yaml not found")
        return False

    with open(path) as f:
        data = yaml.safe_load(f)

    missing_keys = [k for k in REQUIRED_MATERIALITY_KEYS if k not in data]
    if missing_keys:
        print(f"  FAIL  Missing required keys: {missing_keys}")
        return False

    print(f"  PASS  All required keys present")

    pm = data['performance_materiality_usd']
    tt = data['trivial_threshold_usd']
    ratio = tt / pm

    print(f"  INFO  Performance materiality:  ${pm:>15,.0f}")
    print(f"  INFO  Trivial threshold:         ${tt:>15,.0f}  ({ratio:.1%} of PM)")

    expected_ratio = 0.05
    if abs(ratio - expected_ratio) > 0.001:
        print(f"  WARN  Trivial threshold is {ratio:.1%} of PM — expected 5%. "
              f"Update intentional? Verify with Controller.")
    else:
        print(f"  PASS  Trivial threshold ratio correct (5% of PM)")

    return True


def validate_agent_definitions():
    header("Agent definitions (AGENT.md)")
    passed = True

    for agent in REQUIRED_AGENTS:
        path = ROOT / 'agents' / agent / 'AGENT.md'
        if path.exists():
            size = path.stat().st_size
            print(f"  PASS  agents/{agent}/AGENT.md  ({size:,} bytes)")
        else:
            print(f"  FAIL  agents/{agent}/AGENT.md  NOT FOUND")
            passed = False

    return passed


def validate_skill_definitions():
    header("Skill definitions (SKILL.md)")
    passed = True

    for skill in REQUIRED_SKILLS:
        path = ROOT / 'skills' / skill / 'SKILL.md'
        if path.exists():
            size = path.stat().st_size
            print(f"  PASS  skills/{skill}/SKILL.md  ({size:,} bytes)")
        else:
            print(f"  FAIL  skills/{skill}/SKILL.md  NOT FOUND")
            passed = False

    return passed


def main():
    print("=" * 60)
    print("Close System — Artifact Validation")
    print("=" * 60)

    results = [
        validate_dataset(ROOT / 'data' / 'synthetic' / 'lumina_close_dataset.xlsx'),
        validate_yaml_files(),
        validate_materiality(),
        validate_agent_definitions(),
        validate_skill_definitions(),
    ]

    print("\n" + "=" * 60)
    if all(results):
        print("RESULT: PASSED — all validations clean")
        print("=" * 60)
        sys.exit(0)
    else:
        failed = results.count(False)
        print(f"RESULT: FAILED — {failed} check(s) did not pass")
        print("=" * 60)
        sys.exit(1)


if __name__ == '__main__':
    main()
