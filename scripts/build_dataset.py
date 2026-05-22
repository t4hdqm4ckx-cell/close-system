"""Generate Lumina Streaming Co. synthetic close dataset for Nov 2026."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
ARIAL_TITLE = Font(name='Arial', size=14, bold=True)
HDR_FILL = PatternFill('solid', start_color='1F4E78')
SUBHDR_FILL = PatternFill('solid', start_color='DDEBF7')
FLAG_FILL = PatternFill('solid', start_color='FFE699')
TOTAL_FILL = PatternFill('solid', start_color='F2F2F2')
THIN = Side(border_style='thin', color='B4B4B4')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

CURRENCY = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
PCT = '0.0%;(0.0%);"-"'
DATE_FMT = 'yyyy-mm-dd'

wb = Workbook()

# ============================================================
# Sheet 1: README
# ============================================================
ws = wb.active
ws.title = 'README'
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 90

ws['A1'] = 'Lumina Streaming Co. — Month-End Close Synthetic Dataset'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:B1')

rows = [
    ('', ''),
    ('Purpose', 'Synthetic dataset for developing and demonstrating a multi-agent month-end close system. All values are fabricated.'),
    ('Company', 'Lumina Streaming Co. — synthetic streaming/media company, ~$1.8B annual revenue, FYE December.'),
    ('Close period', 'November 2026 (closing in early December 2026)'),
    ('Currency', 'USD, all values in actual dollars (not thousands)'),
    ('Entities', 'LuminaUS, LuminaEMEA, LuminaAPAC'),
    ('', ''),
    ('Sheets', ''),
    ('  TrialBalance', 'GL account-level Nov-26 actual, Oct-26 prior, Nov-26 budget, with entity dimension'),
    ('  AR_Aging', 'Accounts receivable detail by customer, aged in standard buckets'),
    ('  AP_Aging', 'Accounts payable detail by vendor, aged in standard buckets'),
    ('  PrepaidSchedule', 'Prepaid asset detail with amortization schedule'),
    ('  AccrualSchedule', 'Accrual balances with reversal dates'),
    ('  BankRec', 'Bank balance vs book reconciliation for primary operating account'),
    ('  IntercoMatrix', 'Intercompany receivable/payable matrix across the three entities'),
    ('  CloseCalendar', 'Close task list with owners, due dates, dependencies, status'),
    ('  JE_Log', 'November 2026 journal entry log'),
    ('', ''),
    ('Embedded findings', 'Intentional issues designed for agents to surface during demo:'),
    ('  1', 'Bank rec: $45,231 outstanding check (67 days old) — should be voided/investigated'),
    ('  2', 'Intercompany: $12,400 mismatch between LuminaUS receivable from EMEA vs EMEA payable to US'),
    ('  3', 'Accruals: $850,000 Oct-26 marketing accrual was never reversed in Nov-26'),
    ('  4', 'AR aging: Vertex Media has $2,300,000 in the 90+ day bucket'),
    ('  5', 'JE log: JE-2026-11-0042 posted on a Saturday, round $1.5M reclass, no support doc reference'),
    ('  6', 'P&L flux: Performance Marketing is $3.0M unfavorable to budget (-20%) — needs commentary'),
    ('  7', 'P&L flux: CDN/streaming delivery is $0.8M unfavorable (-10%)'),
    ('', ''),
    ('Materiality (Lumina defaults)', ''),
    ('  Performance materiality', '$9,000,000 (≈50bps of annual revenue)'),
    ('  P&L flux trigger', '> $250,000 AND > 5% MoM (both must be true)'),
    ('  BvA flux trigger', '> $500,000 AND > 5% vs budget'),
    ('  BS rec exception', 'Any unreconciled item > $100,000 OR > 30 days old'),
]
for i, (a, b) in enumerate(rows, start=2):
    ws[f'A{i}'] = a
    ws[f'B{i}'] = b
    ws[f'A{i}'].font = ARIAL_BOLD if a and not a.startswith('  ') else ARIAL
    ws[f'B{i}'].font = ARIAL
    ws[f'A{i}'].alignment = LEFT
    ws[f'B{i}'].alignment = LEFT

# ============================================================
# Sheet 2: TrialBalance
# ============================================================
ws = wb.create_sheet('TrialBalance')

tb_headers = ['Account', 'Account Name', 'Statement', 'Entity', 'Nov-26 Actual', 'Oct-26 Prior', 'Nov-26 Budget', 'MoM Variance', 'BvA Variance', 'BvA %']
for c, h in enumerate(tb_headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = ARIAL_HDR
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = BOX

# Account data: (account, name, statement, entity, nov_actual, oct_prior, nov_budget)
# Note: P&L accounts: positive = debit (expense), negative = credit (revenue)
# Note: positive amounts only; conventions: revenue shown as positive, expenses positive, contra accounts negative
tb_data = [
    # ASSETS
    ('100100', 'Cash - Operating', 'BS', 'LuminaUS', 142500000, 138900000, 140000000),
    ('100100', 'Cash - Operating', 'BS', 'LuminaEMEA', 38200000, 36500000, 37000000),
    ('100100', 'Cash - Operating', 'BS', 'LuminaAPAC', 22100000, 21800000, 21500000),
    ('100200', 'Cash - Money Market', 'BS', 'LuminaUS', 285000000, 280000000, 285000000),
    ('110100', 'AR - Subscriptions', 'BS', 'LuminaUS', 48200000, 47100000, 46500000),
    ('110100', 'AR - Subscriptions', 'BS', 'LuminaEMEA', 18900000, 17800000, 18200000),
    ('110100', 'AR - Subscriptions', 'BS', 'LuminaAPAC', 9400000, 9100000, 9200000),
    ('110200', 'AR - Advertising', 'BS', 'LuminaUS', 31500000, 28200000, 29000000),
    ('110300', 'AR - Other', 'BS', 'LuminaUS', 2400000, 2100000, 2000000),
    ('120100', 'Prepaid Content Licenses', 'BS', 'LuminaUS', 215000000, 248000000, 220000000),
    ('120200', 'Prepaid Software', 'BS', 'LuminaUS', 8400000, 9200000, 8500000),
    ('120300', 'Prepaid Insurance', 'BS', 'LuminaUS', 3200000, 3500000, 3300000),
    ('120400', 'Other Prepaids', 'BS', 'LuminaUS', 1800000, 1900000, 1900000),
    ('130100', 'IC Receivable - LuminaEMEA', 'BS', 'LuminaUS', 12400000, 11800000, 12000000),  # mismatch with EMEA payable
    ('130200', 'IC Receivable - LuminaAPAC', 'BS', 'LuminaUS', 8700000, 8500000, 8600000),
    ('140100', 'Capitalized Content - Net', 'BS', 'LuminaUS', 1850000000, 1885000000, 1855000000),
    ('150100', 'PP&E - Net', 'BS', 'LuminaUS', 142000000, 144500000, 143000000),
    ('160100', 'Goodwill', 'BS', 'LuminaUS', 412000000, 412000000, 412000000),
    ('170100', 'Other Assets', 'BS', 'LuminaUS', 28500000, 28200000, 28000000),
    # LIABILITIES
    ('200100', 'AP - Trade', 'BS', 'LuminaUS', 38500000, 41200000, 39000000),
    ('200200', 'AP - Content Vendors', 'BS', 'LuminaUS', 87200000, 92500000, 88000000),
    ('210100', 'Accrued Bonus', 'BS', 'LuminaUS', 28400000, 26000000, 28000000),
    ('210200', 'Accrued Vacation', 'BS', 'LuminaUS', 12500000, 12100000, 12300000),
    ('210300', 'Accrued Content Costs', 'BS', 'LuminaUS', 45000000, 48500000, 44000000),
    ('210400', 'Accrued Marketing', 'BS', 'LuminaUS', 14800000, 8200000, 6500000),  # contains the stale $850K from Oct
    ('210500', 'Other Accruals', 'BS', 'LuminaUS', 8900000, 9100000, 8800000),
    ('220100', 'Deferred Revenue', 'BS', 'LuminaUS', 124000000, 119000000, 121000000),
    ('230100', 'IC Payable - LuminaUS', 'BS', 'LuminaEMEA', 0, 11800000, 12000000),  # MISMATCH: should be 12,400,000
    ('230200', 'IC Payable - LuminaUS', 'BS', 'LuminaAPAC', 8700000, 8500000, 8600000),
    ('240100', 'Long-term Debt', 'BS', 'LuminaUS', 750000000, 750000000, 750000000),
    ('250100', 'Other Liabilities', 'BS', 'LuminaUS', 42000000, 41500000, 42000000),
    # EQUITY
    ('300100', 'Common Stock', 'BS', 'LuminaUS', 1000000, 1000000, 1000000),
    ('300200', 'Additional Paid-in Capital', 'BS', 'LuminaUS', 1850000000, 1850000000, 1850000000),
    ('300300', 'Retained Earnings', 'BS', 'LuminaUS', 425000000, 425000000, 425000000),
    # REVENUE (positive = revenue earned)
    ('400100', 'Subscription Revenue', 'IS', 'LuminaUS', 74200000, 73500000, 72500000),
    ('400100', 'Subscription Revenue', 'IS', 'LuminaEMEA', 22800000, 22100000, 22000000),
    ('400100', 'Subscription Revenue', 'IS', 'LuminaAPAC', 9000000, 8800000, 9500000),
    ('400200', 'Advertising Revenue', 'IS', 'LuminaUS', 32500000, 28200000, 31000000),
    ('400200', 'Advertising Revenue', 'IS', 'LuminaEMEA', 8200000, 7400000, 8000000),
    ('400200', 'Advertising Revenue', 'IS', 'LuminaAPAC', 3300000, 3000000, 4000000),
    ('400300', 'Other Revenue', 'IS', 'LuminaUS', 2000000, 1900000, 2000000),
    # COGS
    ('500100', 'Content Amortization', 'IS', 'LuminaUS', 32500000, 31800000, 32000000),
    ('500200', 'Streaming Delivery (CDN)', 'IS', 'LuminaUS', 9000000, 8400000, 8200000),  # $0.8M over (~10%)
    ('500300', 'Royalties', 'IS', 'LuminaUS', 3500000, 3400000, 3500000),
    ('500400', 'Other COGS', 'IS', 'LuminaUS', 1100000, 1050000, 1100000),
    # OPEX
    ('600100', 'Marketing - Performance', 'IS', 'LuminaUS', 18000000, 14500000, 15000000),  # $3.0M over (-20%)
    ('600200', 'Marketing - Brand', 'IS', 'LuminaUS', 7000000, 6800000, 7000000),
    ('600300', 'Sales Commissions', 'IS', 'LuminaUS', 3500000, 3200000, 3400000),
    ('600400', 'Salaries - Engineering', 'IS', 'LuminaUS', 14000000, 14100000, 14500000),
    ('600500', 'Salaries - G&A', 'IS', 'LuminaUS', 7500000, 7400000, 7500000),
    ('600600', 'Tech Infrastructure', 'IS', 'LuminaUS', 5000000, 4900000, 5000000),
    ('600700', 'Professional Fees', 'IS', 'LuminaUS', 1200000, 1100000, 1000000),
    ('600800', 'Travel & Entertainment', 'IS', 'LuminaUS', 800000, 750000, 900000),
    ('600900', 'Office & Facilities', 'IS', 'LuminaUS', 1500000, 1500000, 1500000),
    ('601000', 'Other OpEx', 'IS', 'LuminaUS', 1500000, 1450000, 1500000),
    # NON-OP
    ('700100', 'Interest Income', 'IS', 'LuminaUS', -800000, -750000, -750000),
    ('700200', 'Interest Expense', 'IS', 'LuminaUS', 3200000, 3200000, 3200000),
    ('700300', 'FX Gain/Loss', 'IS', 'LuminaUS', 450000, -200000, 0),
    # TAX
    ('800100', 'Income Tax Expense', 'IS', 'LuminaUS', 6800000, 6500000, 7000000),
]

for r, row in enumerate(tb_data, start=2):
    acct, name, stmt, entity, actual, prior, budget = row
    ws.cell(row=r, column=1, value=acct).font = ARIAL
    ws.cell(row=r, column=2, value=name).font = ARIAL
    ws.cell(row=r, column=3, value=stmt).font = ARIAL
    ws.cell(row=r, column=4, value=entity).font = ARIAL
    ws.cell(row=r, column=5, value=actual).font = ARIAL
    ws.cell(row=r, column=5).number_format = CURRENCY
    ws.cell(row=r, column=6, value=prior).font = ARIAL
    ws.cell(row=r, column=6).number_format = CURRENCY
    ws.cell(row=r, column=7, value=budget).font = ARIAL
    ws.cell(row=r, column=7).number_format = CURRENCY
    # Formulas for variances
    ws.cell(row=r, column=8, value=f'=E{r}-F{r}').number_format = CURRENCY
    ws.cell(row=r, column=9, value=f'=E{r}-G{r}').number_format = CURRENCY
    ws.cell(row=r, column=10, value=f'=IFERROR((E{r}-G{r})/G{r},0)').number_format = PCT
    for c in range(1, 11):
        ws.cell(row=r, column=c).border = BOX

col_widths = [11, 32, 11, 14, 16, 16, 16, 16, 16, 11]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'

# ============================================================
# Sheet 3: AR_Aging
# ============================================================
ws = wb.create_sheet('AR_Aging')
ar_headers = ['Customer ID', 'Customer Name', 'AR Type', 'Entity', 'Invoice Date', 'Due Date', 'Original Amount', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', '90+ Days', 'Total']
for c, h in enumerate(ar_headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

ar_data = [
    ('CUST-1001', 'Vertex Media', 'Advertising', 'LuminaUS', date(2026, 7, 15), date(2026, 8, 14), 2300000, 0, 0, 0, 0, 2300000),  # FLAG
    ('CUST-1002', 'Hudson Brands Group', 'Advertising', 'LuminaUS', date(2026, 11, 5), date(2026, 12, 5), 4500000, 4500000, 0, 0, 0, 0),
    ('CUST-1003', 'Pacific Auto Co.', 'Advertising', 'LuminaUS', date(2026, 10, 12), date(2026, 11, 11), 3200000, 0, 3200000, 0, 0, 0),
    ('CUST-1004', 'Greentide Telecom', 'Advertising', 'LuminaUS', date(2026, 9, 20), date(2026, 10, 20), 1800000, 0, 0, 1800000, 0, 0),
    ('CUST-1005', 'Norrland Foods', 'Advertising', 'LuminaEMEA', date(2026, 11, 1), date(2026, 12, 1), 2900000, 2900000, 0, 0, 0, 0),
    ('CUST-1006', 'Atlas Insurance', 'Advertising', 'LuminaUS', date(2026, 8, 28), date(2026, 9, 27), 1200000, 0, 0, 0, 1200000, 0),
    ('CUST-1007', 'Sterling Retail', 'Advertising', 'LuminaUS', date(2026, 11, 12), date(2026, 12, 12), 3800000, 3800000, 0, 0, 0, 0),
    ('CUST-1008', 'Bayline Financial', 'Advertising', 'LuminaUS', date(2026, 10, 28), date(2026, 11, 27), 2700000, 0, 2700000, 0, 0, 0),
    ('CUST-1009', 'Westcoast Mobility', 'Advertising', 'LuminaUS', date(2026, 11, 18), date(2026, 12, 18), 5100000, 5100000, 0, 0, 0, 0),
    ('CUST-1010', 'Aurora Beverage', 'Advertising', 'LuminaUS', date(2026, 9, 10), date(2026, 10, 10), 2100000, 0, 0, 2100000, 0, 0),
    ('CUST-2001', 'B2B Subscriber Block - US', 'Subscriptions', 'LuminaUS', date(2026, 11, 1), date(2026, 11, 30), 48200000, 48200000, 0, 0, 0, 0),
    ('CUST-2002', 'B2B Subscriber Block - EMEA', 'Subscriptions', 'LuminaEMEA', date(2026, 11, 1), date(2026, 11, 30), 18900000, 18900000, 0, 0, 0, 0),
    ('CUST-2003', 'B2B Subscriber Block - APAC', 'Subscriptions', 'LuminaAPAC', date(2026, 11, 1), date(2026, 11, 30), 9400000, 9400000, 0, 0, 0, 0),
    ('CUST-3001', 'Misc Receivables Block', 'Other', 'LuminaUS', date(2026, 10, 15), date(2026, 11, 14), 2400000, 0, 2400000, 0, 0, 0),
]
for r, row in enumerate(ar_data, start=2):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c >= 7:
            cell.number_format = CURRENCY
        if c in (5, 6):
            cell.number_format = DATE_FMT
    # Total column M = sum of aging buckets H:L
    ws.cell(row=r, column=13, value=f'=SUM(H{r}:L{r})').font = ARIAL
    ws.cell(row=r, column=13).number_format = CURRENCY
    ws.cell(row=r, column=13).border = BOX
    # Flag Vertex Media row
    if row[0] == 'CUST-1001':
        for c in range(1, 14):
            ws.cell(row=r, column=c).fill = FLAG_FILL

# Totals row
total_row = len(ar_data) + 2
ws.cell(row=total_row, column=1, value='TOTAL').font = ARIAL_BOLD
for c in range(7, 14):
    col = get_column_letter(c)
    ws.cell(row=total_row, column=c, value=f'=SUM({col}2:{col}{total_row-1})').font = ARIAL_BOLD
    ws.cell(row=total_row, column=c).number_format = CURRENCY
    ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=c).border = BOX
ws.cell(row=total_row, column=1).fill = TOTAL_FILL

for i, w in enumerate([12, 28, 14, 12, 13, 13, 16, 14, 14, 14, 14, 14, 16], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ============================================================
# Sheet 4: AP_Aging
# ============================================================
ws = wb.create_sheet('AP_Aging')
ap_headers = ['Vendor ID', 'Vendor Name', 'Category', 'Entity', 'Invoice Date', 'Due Date', 'Original Amount', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', '90+ Days', 'Total']
for c, h in enumerate(ap_headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

ap_data = [
    ('VEND-5001', 'Pinnacle Studios LLC', 'Content', 'LuminaUS', date(2026, 11, 15), date(2026, 12, 15), 28000000, 28000000, 0, 0, 0, 0),
    ('VEND-5002', 'Northgate Pictures', 'Content', 'LuminaUS', date(2026, 10, 30), date(2026, 11, 29), 22000000, 0, 22000000, 0, 0, 0),
    ('VEND-5003', 'Skylark Productions', 'Content', 'LuminaUS', date(2026, 11, 20), date(2026, 12, 20), 18500000, 18500000, 0, 0, 0, 0),
    ('VEND-5004', 'Helix Films', 'Content', 'LuminaUS', date(2026, 11, 8), date(2026, 12, 8), 12200000, 12200000, 0, 0, 0, 0),
    ('VEND-5005', 'Cascade Media Partners', 'Content', 'LuminaUS', date(2026, 10, 22), date(2026, 11, 21), 6500000, 0, 6500000, 0, 0, 0),
    ('VEND-6001', 'AWS', 'Tech Infra', 'LuminaUS', date(2026, 11, 25), date(2026, 12, 25), 5000000, 5000000, 0, 0, 0, 0),
    ('VEND-6002', 'Akamai', 'Tech Infra', 'LuminaUS', date(2026, 11, 18), date(2026, 12, 18), 4500000, 4500000, 0, 0, 0, 0),
    ('VEND-7001', 'Meta Ads', 'Marketing', 'LuminaUS', date(2026, 11, 30), date(2026, 12, 30), 8000000, 8000000, 0, 0, 0, 0),
    ('VEND-7002', 'Google Ads', 'Marketing', 'LuminaUS', date(2026, 11, 30), date(2026, 12, 30), 7500000, 7500000, 0, 0, 0, 0),
    ('VEND-7003', 'TikTok Ads', 'Marketing', 'LuminaUS', date(2026, 11, 30), date(2026, 12, 30), 2500000, 2500000, 0, 0, 0, 0),
    ('VEND-8001', 'Deloitte', 'Professional', 'LuminaUS', date(2026, 11, 12), date(2026, 12, 12), 800000, 800000, 0, 0, 0, 0),
    ('VEND-8002', 'Latham & Watkins', 'Professional', 'LuminaUS', date(2026, 11, 5), date(2026, 12, 5), 400000, 400000, 0, 0, 0, 0),
    ('VEND-9001', 'Trade AP - LuminaEMEA Block', 'Trade', 'LuminaEMEA', date(2026, 11, 1), date(2026, 12, 1), 18000000, 18000000, 0, 0, 0, 0),
    ('VEND-9002', 'Trade AP - LuminaAPAC Block', 'Trade', 'LuminaAPAC', date(2026, 11, 1), date(2026, 12, 1), 9800000, 9800000, 0, 0, 0, 0),
]
for r, row in enumerate(ap_data, start=2):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c >= 7:
            cell.number_format = CURRENCY
        if c in (5, 6):
            cell.number_format = DATE_FMT
    # Total column M = sum of aging buckets H:L
    ws.cell(row=r, column=13, value=f'=SUM(H{r}:L{r})').font = ARIAL
    ws.cell(row=r, column=13).number_format = CURRENCY
    ws.cell(row=r, column=13).border = BOX

total_row = len(ap_data) + 2
ws.cell(row=total_row, column=1, value='TOTAL').font = ARIAL_BOLD
for c in range(7, 14):
    col = get_column_letter(c)
    ws.cell(row=total_row, column=c, value=f'=SUM({col}2:{col}{total_row-1})').font = ARIAL_BOLD
    ws.cell(row=total_row, column=c).number_format = CURRENCY
    ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=c).border = BOX
ws.cell(row=total_row, column=1).fill = TOTAL_FILL

for i, w in enumerate([12, 30, 13, 12, 13, 13, 16, 14, 14, 14, 14, 14, 16], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ============================================================
# Sheet 5: PrepaidSchedule
# ============================================================
ws = wb.create_sheet('PrepaidSchedule')
pp_headers = ['Asset ID', 'Description', 'Category', 'Vendor', 'Start Date', 'End Date', 'Total Cost', 'Months', 'Monthly Amort', 'Oct-26 Balance', 'Nov-26 Amort', 'Nov-26 Balance']
for c, h in enumerate(pp_headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

pp_data = [
    ('PP-2026-001', 'Content license — Pinnacle Tier 1 slate', 'Content License', 'Pinnacle Studios', date(2026, 1, 1), date(2026, 12, 31), 120000000, 12, date(2026, 6, 1)),
    ('PP-2026-002', 'Content license — Northgate exclusive', 'Content License', 'Northgate Pictures', date(2026, 4, 1), date(2027, 3, 31), 84000000, 12, date(2026, 6, 1)),
    ('PP-2026-003', 'Content license — Skylark catalog refresh', 'Content License', 'Skylark Productions', date(2026, 7, 1), date(2027, 6, 30), 36000000, 12, date(2026, 6, 1)),
    ('PP-2025-018', 'Content license — Helix bundle (prior year carry)', 'Content License', 'Helix Films', date(2025, 10, 1), date(2026, 9, 30), 24000000, 12, date(2026, 6, 1)),
    ('PP-2026-004', 'Salesforce annual subscription', 'Software', 'Salesforce', date(2026, 1, 1), date(2026, 12, 31), 4800000, 12, date(2026, 6, 1)),
    ('PP-2026-005', 'Workday annual subscription', 'Software', 'Workday', date(2026, 1, 1), date(2026, 12, 31), 3600000, 12, date(2026, 6, 1)),
    ('PP-2026-006', 'D&O insurance', 'Insurance', 'Aon', date(2026, 7, 1), date(2027, 6, 30), 2400000, 12, date(2026, 6, 1)),
    ('PP-2026-007', 'General liability insurance', 'Insurance', 'Marsh', date(2026, 7, 1), date(2027, 6, 30), 1200000, 12, date(2026, 6, 1)),
    ('PP-2026-008', 'Industry conference sponsorships', 'Other', 'Various', date(2026, 9, 1), date(2027, 8, 31), 2400000, 12, date(2026, 6, 1)),
]

for r, row in enumerate(pp_data, start=2):
    asset_id, desc, cat, vendor, start, end, total, months, _ = row
    ws.cell(row=r, column=1, value=asset_id).font = ARIAL
    ws.cell(row=r, column=2, value=desc).font = ARIAL
    ws.cell(row=r, column=3, value=cat).font = ARIAL
    ws.cell(row=r, column=4, value=vendor).font = ARIAL
    ws.cell(row=r, column=5, value=start).font = ARIAL
    ws.cell(row=r, column=5).number_format = DATE_FMT
    ws.cell(row=r, column=6, value=end).font = ARIAL
    ws.cell(row=r, column=6).number_format = DATE_FMT
    ws.cell(row=r, column=7, value=total).font = ARIAL
    ws.cell(row=r, column=7).number_format = CURRENCY
    ws.cell(row=r, column=8, value=months).font = ARIAL
    ws.cell(row=r, column=8).alignment = CENTER
    # Monthly amort = total/months
    ws.cell(row=r, column=9, value=f'=G{r}/H{r}').number_format = CURRENCY
    # Oct-26 balance (hardcoded to make a realistic snapshot)
    # Nov-26 amort = monthly amort
    # Nov-26 balance = Oct balance - Nov amort
    pass

# Hardcoded Oct-26 balances and let formulas compute Nov
oct_balances = [10000000, 56000000, 33000000, 0, 400000, 300000, 2200000, 1100000, 2200000]
for i, ob in enumerate(oct_balances):
    r = i + 2
    ws.cell(row=r, column=10, value=ob).number_format = CURRENCY
    ws.cell(row=r, column=11, value=f'=MIN(I{r},J{r})').number_format = CURRENCY  # don't amortize more than balance
    ws.cell(row=r, column=12, value=f'=J{r}-K{r}').number_format = CURRENCY
    for c in range(1, 13):
        ws.cell(row=r, column=c).border = BOX

total_row = len(pp_data) + 2
ws.cell(row=total_row, column=1, value='TOTAL').font = ARIAL_BOLD
ws.cell(row=total_row, column=1).fill = TOTAL_FILL
for c in (7, 9, 10, 11, 12):
    col = get_column_letter(c)
    ws.cell(row=total_row, column=c, value=f'=SUM({col}2:{col}{total_row-1})').font = ARIAL_BOLD
    ws.cell(row=total_row, column=c).number_format = CURRENCY
    ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=c).border = BOX

for i, w in enumerate([13, 38, 16, 22, 12, 12, 16, 9, 15, 16, 15, 16], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ============================================================
# Sheet 6: AccrualSchedule
# ============================================================
ws = wb.create_sheet('AccrualSchedule')
ac_headers = ['Accrual ID', 'Description', 'Category', 'Period Originated', 'Reversal Status', 'Oct-26 Balance', 'Nov-26 Activity', 'Nov-26 Balance', 'Notes']
for c, h in enumerate(ac_headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

ac_data = [
    ('ACR-2026-11-01', 'Q4 bonus accrual — corporate', 'Bonus', 'Nov-26', 'Active', 26000000, 2400000, 28400000, 'Trues up to FY target; OK'),
    ('ACR-2026-11-02', 'Vacation accrual — true-up', 'Vacation', 'Nov-26', 'Active', 12100000, 400000, 12500000, 'Standard monthly true-up'),
    ('ACR-2026-11-03', 'Content costs not yet invoiced — Pinnacle', 'Content', 'Nov-26', 'Reverses Dec-26', 18000000, 4000000, 22000000, 'OK'),
    ('ACR-2026-11-04', 'Content costs not yet invoiced — Northgate', 'Content', 'Nov-26', 'Reverses Dec-26', 15000000, 1500000, 16500000, 'OK'),
    ('ACR-2026-11-05', 'Content costs not yet invoiced — Other', 'Content', 'Nov-26', 'Reverses Dec-26', 15500000, -9000000, 6500000, 'Several invoices received in Nov'),
    ('ACR-2026-10-12', 'Performance marketing — Meta retargeting', 'Marketing', 'Oct-26', 'NOT REVERSED', 850000, 0, 850000, 'STALE: should have reversed in Nov when invoice posted'),  # FLAG
    ('ACR-2026-11-06', 'Performance marketing — Nov accrual', 'Marketing', 'Nov-26', 'Reverses Dec-26', 5500000, 8450000, 13950000, 'High due to Nov campaign spend'),
    ('ACR-2026-11-07', 'Sales commission accrual', 'Compensation', 'Nov-26', 'Reverses Dec-26', 1850000, 200000, 2050000, 'OK'),
    ('ACR-2026-11-08', 'Professional fees — legal', 'Professional', 'Nov-26', 'Reverses Dec-26', 450000, 50000, 500000, 'OK'),
    ('ACR-2026-11-09', 'Utilities and facilities', 'Other', 'Nov-26', 'Reverses Dec-26', 850000, 200000, 1050000, 'OK'),
    ('ACR-2026-11-10', 'Other operating accruals — block', 'Other', 'Nov-26', 'Reverses Dec-26', 6950000, 400000, 7350000, 'OK'),
]
for r, row in enumerate(ac_data, start=2):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c in (6, 7, 8):
            cell.number_format = CURRENCY
        if c == 9:
            cell.alignment = LEFT
    if row[0] == 'ACR-2026-10-12':
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = FLAG_FILL

total_row = len(ac_data) + 2
ws.cell(row=total_row, column=1, value='TOTAL').font = ARIAL_BOLD
ws.cell(row=total_row, column=1).fill = TOTAL_FILL
for c in (6, 7, 8):
    col = get_column_letter(c)
    ws.cell(row=total_row, column=c, value=f'=SUM({col}2:{col}{total_row-1})').font = ARIAL_BOLD
    ws.cell(row=total_row, column=c).number_format = CURRENCY
    ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=c).border = BOX

for i, w in enumerate([16, 44, 14, 16, 18, 16, 16, 16, 50], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ============================================================
# Sheet 7: BankRec
# ============================================================
ws = wb.create_sheet('BankRec')
ws['A1'] = 'Bank Reconciliation — LuminaUS Operating Account #4471'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:D1')
ws['A2'] = 'Period: November 30, 2026'
ws['A2'].font = ARIAL_BOLD
ws.merge_cells('A2:D2')

# Summary table
ws['A4'] = 'Reconciliation Summary'
ws['A4'].font = ARIAL_BOLD
ws['A4'].fill = SUBHDR_FILL
ws.merge_cells('A4:D4')

summary = [
    ('Bank statement balance — Nov 30, 2026', 144855700),
    ('Add: Deposits in transit', 3200000),
    ('Less: Outstanding checks (see detail)', -5550000),
    ('Adjusted bank balance', None),  # formula
    ('', None),
    ('Book balance — GL 100100 LuminaUS', 142500000),
    ('Reconciling items (per detail)', None),  # formula
    ('Adjusted book balance', None),  # formula
    ('', None),
    ('DIFFERENCE (should be zero)', None),
]
for i, (lbl, val) in enumerate(summary):
    r = i + 5
    ws.cell(row=r, column=1, value=lbl).font = ARIAL_BOLD if lbl else ARIAL
    if val is not None:
        ws.cell(row=r, column=2, value=val).number_format = CURRENCY
        ws.cell(row=r, column=2).font = ARIAL

# Formulas
ws['B8'] = '=B5+B6+B7'  # adjusted bank
ws['B11'] = '=SUMIF(C20:C40,"Book",E20:E40)'  # reconciling items
ws['B12'] = '=B10+B11'  # adjusted book
ws['B14'] = '=B8-B12'  # difference
ws['B14'].font = ARIAL_BOLD
ws['B14'].fill = FLAG_FILL  # if nonzero this flags
for r in (8, 11, 12, 14):
    ws.cell(row=r, column=2).number_format = CURRENCY
    ws.cell(row=r, column=2).font = ARIAL_BOLD
    ws.cell(row=r, column=1).font = ARIAL_BOLD

# Outstanding checks detail
ws['A17'] = 'Outstanding Checks Detail'
ws['A17'].font = ARIAL_BOLD
ws['A17'].fill = SUBHDR_FILL
ws.merge_cells('A17:E17')

check_headers = ['Check #', 'Date Issued', 'Side (Bank/Book)', 'Payee', 'Amount']
for c, h in enumerate(check_headers, start=1):
    cell = ws.cell(row=19, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

checks = [
    ('4518', date(2026, 11, 24), 'Bank', 'Pinnacle Studios — content fee', 2800000),
    ('4519', date(2026, 11, 26), 'Bank', 'AWS — Nov hosting', 1200000),
    ('4520', date(2026, 11, 28), 'Bank', 'Payroll — Nov supplemental', 1505000),
    ('4521', date(2026, 9, 24), 'Bank', 'Greenway Productions — content invoice', 45000),  # FLAG: 67 days old
    ('DIT-1', date(2026, 11, 30), 'Bank', 'Deposit in transit — adjustment', 0),
    # Book-side reconciling items
    ('BK-1', date(2026, 11, 30), 'Book', 'Wire fee not yet recorded', -2500),
    ('BK-2', date(2026, 11, 30), 'Book', 'Interest credit not yet recorded', 8200),
]
for i, row in enumerate(checks, start=20):
    chk, dt, side, payee, amt = row
    if chk == 'DIT-1' or chk.startswith('BK'):
        continue
    ws.cell(row=i, column=1, value=chk).font = ARIAL
    ws.cell(row=i, column=2, value=dt).font = ARIAL
    ws.cell(row=i, column=2).number_format = DATE_FMT
    ws.cell(row=i, column=3, value=side).font = ARIAL
    ws.cell(row=i, column=4, value=payee).font = ARIAL
    ws.cell(row=i, column=5, value=amt).font = ARIAL
    ws.cell(row=i, column=5).number_format = CURRENCY
    if chk == '4521':
        for c in range(1, 6):
            ws.cell(row=i, column=c).fill = FLAG_FILL
    for c in range(1, 6):
        ws.cell(row=i, column=c).border = BOX

# Book-side reconciling items at row 25+
ws.cell(row=25, column=1, value='BK-001').font = ARIAL
ws.cell(row=25, column=2, value=date(2026, 11, 30)).number_format = DATE_FMT
ws.cell(row=25, column=3, value='Book').font = ARIAL
ws.cell(row=25, column=4, value='Wire fees not yet recorded in GL').font = ARIAL
ws.cell(row=25, column=5, value=-2500).number_format = CURRENCY

ws.cell(row=26, column=1, value='BK-002').font = ARIAL
ws.cell(row=26, column=2, value=date(2026, 11, 30)).number_format = DATE_FMT
ws.cell(row=26, column=3, value='Book').font = ARIAL
ws.cell(row=26, column=4, value='Interest credit not yet recorded in GL').font = ARIAL
ws.cell(row=26, column=5, value=8200).number_format = CURRENCY

for r in (25, 26):
    for c in range(1, 6):
        ws.cell(row=r, column=c).font = ARIAL
        ws.cell(row=r, column=c).border = BOX

for i, w in enumerate([42, 16, 18, 38, 16], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 8: IntercoMatrix
# ============================================================
ws = wb.create_sheet('IntercoMatrix')
ws['A1'] = 'Intercompany Balance Matrix — November 30, 2026'
ws['A1'].font = ARIAL_TITLE
ws.merge_cells('A1:E1')
ws['A2'] = 'Each row shows balance booked BY the entity in column A AGAINST the counterparty in column B'
ws['A2'].font = ARIAL
ws.merge_cells('A2:E2')

ic_headers = ['Entity (booker)', 'Counterparty', 'Balance Type', 'Amount', 'Match Status']
for c, h in enumerate(ic_headers, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

ic_data = [
    ('LuminaUS', 'LuminaEMEA', 'Receivable', 12400000, 'CHECK'),  # mismatch
    ('LuminaEMEA', 'LuminaUS', 'Payable', 0, 'CHECK'),
    ('LuminaUS', 'LuminaAPAC', 'Receivable', 8700000, 'OK'),
    ('LuminaAPAC', 'LuminaUS', 'Payable', 8700000, 'OK'),
    ('LuminaEMEA', 'LuminaAPAC', 'Receivable', 2100000, 'OK'),
    ('LuminaAPAC', 'LuminaEMEA', 'Payable', 2100000, 'OK'),
]
for i, row in enumerate(ic_data, start=5):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c == 4:
            cell.number_format = CURRENCY
    if row[4] == 'CHECK':
        for c in range(1, 6):
            ws.cell(row=i, column=c).fill = FLAG_FILL

ws['A12'] = 'Validation'
ws['A12'].font = ARIAL_BOLD
ws['A12'].fill = SUBHDR_FILL
ws.merge_cells('A12:E12')
ws['A13'] = 'US→EMEA receivable vs EMEA→US payable'
ws['B13'] = '=D5-D6'
ws['B13'].number_format = CURRENCY
ws['B13'].font = ARIAL_BOLD
ws['B13'].fill = FLAG_FILL
ws['C13'] = 'MISMATCH — $12.4M difference'
ws['C13'].font = ARIAL_BOLD

ws['A14'] = 'US→APAC receivable vs APAC→US payable'
ws['B14'] = '=D7-D8'
ws['B14'].number_format = CURRENCY
ws['B14'].font = ARIAL_BOLD
ws['C14'] = 'OK'

ws['A15'] = 'EMEA→APAC receivable vs APAC→EMEA payable'
ws['B15'] = '=D9-D10'
ws['B15'].number_format = CURRENCY
ws['B15'].font = ARIAL_BOLD
ws['C15'] = 'OK'

for i, w in enumerate([18, 18, 14, 18, 35], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 9: CloseCalendar
# ============================================================
ws = wb.create_sheet('CloseCalendar')
cal_headers = ['Task ID', 'Task', 'Phase', 'Owner', 'Business Day', 'Target Date', 'Dependencies', 'Status', 'Notes']
for c, h in enumerate(cal_headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

cal_data = [
    ('T-001', 'Period cutoff confirmation', 'BD1', 'Controller', 'BD1', date(2026, 12, 1), '', 'Complete', ''),
    ('T-002', 'Sub-ledger feeds loaded (AR/AP)', 'BD1', 'Sr Accountant — Sub-ledger', 'BD1', date(2026, 12, 1), 'T-001', 'Complete', ''),
    ('T-003', 'Bank statement received and loaded', 'BD1', 'Sr Accountant — Treasury', 'BD1', date(2026, 12, 1), 'T-001', 'Complete', ''),
    ('T-004', 'Bank reconciliation — operating account', 'BD2', 'Sr Accountant — Treasury', 'BD2', date(2026, 12, 2), 'T-003', 'In Progress', 'Outstanding check #4521 flagged'),
    ('T-005', 'Intercompany matching', 'BD2', 'Sr Accountant — Consolidations', 'BD2', date(2026, 12, 2), 'T-002', 'In Progress', 'US-EMEA mismatch under investigation'),
    ('T-006', 'Standard monthly accruals', 'BD2', 'Sr Accountant — Accruals', 'BD2', date(2026, 12, 2), 'T-001', 'In Progress', ''),
    ('T-007', 'Prepaid amortization JEs', 'BD3', 'Sr Accountant — GL', 'BD3', date(2026, 12, 3), 'T-001', 'Not Started', ''),
    ('T-008', 'Content amortization JEs', 'BD3', 'Content Accounting Lead', 'BD3', date(2026, 12, 3), 'T-001', 'Not Started', ''),
    ('T-009', 'JE review — pass 1', 'BD3', 'Assistant Controller', 'BD3', date(2026, 12, 3), 'T-004,T-005,T-006,T-007,T-008', 'Not Started', ''),
    ('T-010', 'Flux analysis — P&L', 'BD4', 'FP&A Manager', 'BD4', date(2026, 12, 4), 'T-009', 'Not Started', ''),
    ('T-011', 'TB review and adjusting entries', 'BD4', 'Controller', 'BD4', date(2026, 12, 4), 'T-009', 'Not Started', ''),
    ('T-012', 'Balance sheet flux review', 'BD4', 'FP&A Manager', 'BD4', date(2026, 12, 4), 'T-009', 'Not Started', ''),
    ('T-013', 'Final TB locked', 'BD5', 'Controller', 'BD5', date(2026, 12, 5), 'T-011', 'Not Started', ''),
    ('T-014', 'Close package assembled', 'BD5', 'FP&A Manager', 'BD5', date(2026, 12, 5), 'T-013,T-010,T-012', 'Not Started', ''),
    ('T-015', 'CFO review and sign-off', 'BD5', 'CFO', 'BD5', date(2026, 12, 5), 'T-014', 'Not Started', ''),
    ('T-016', 'Reporting submission (HFM/EPM)', 'BD5', 'Sr Accountant — Reporting', 'BD5', date(2026, 12, 5), 'T-015', 'Not Started', ''),
]
for r, row in enumerate(cal_data, start=2):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c == 6:
            cell.number_format = DATE_FMT
        if c in (2, 7, 9):
            cell.alignment = LEFT
    status = row[7]
    if status == 'Complete':
        ws.cell(row=r, column=8).fill = PatternFill('solid', start_color='C6EFCE')
    elif status == 'In Progress':
        ws.cell(row=r, column=8).fill = PatternFill('solid', start_color='FFEB9C')
    else:
        ws.cell(row=r, column=8).fill = PatternFill('solid', start_color='F2F2F2')

for i, w in enumerate([10, 42, 8, 32, 13, 13, 24, 14, 40], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ============================================================
# Sheet 10: JE_Log
# ============================================================
ws = wb.create_sheet('JE_Log')
je_headers = ['JE #', 'Post Date', 'Day of Week', 'Posted By', 'Description', 'Debit Account', 'Credit Account', 'Amount', 'Support Doc Ref', 'Flag']
for c, h in enumerate(je_headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = ARIAL_HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BOX

je_data = [
    ('JE-2026-11-0035', date(2026, 11, 30), 'Monday', 'Sarah Chen', 'Prepaid content amortization — Pinnacle', '500100 Content Amortization', '120100 Prepaid Content Licenses', 10000000, 'AMORT-2026-11-001', ''),
    ('JE-2026-11-0036', date(2026, 11, 30), 'Monday', 'Sarah Chen', 'Prepaid content amortization — Northgate', '500100 Content Amortization', '120100 Prepaid Content Licenses', 7000000, 'AMORT-2026-11-002', ''),
    ('JE-2026-11-0037', date(2026, 11, 30), 'Monday', 'Sarah Chen', 'Prepaid software amortization', '600600 Tech Infrastructure', '120200 Prepaid Software', 700000, 'AMORT-2026-11-003', ''),
    ('JE-2026-11-0038', date(2026, 11, 30), 'Monday', 'David Park', 'Capitalized content amortization — net', '500100 Content Amortization', '140100 Capitalized Content - Net', 15500000, 'AMORT-2026-11-004', ''),
    ('JE-2026-11-0039', date(2026, 11, 30), 'Monday', 'Maria Lopez', 'Q4 bonus accrual true-up', '600500 Salaries - G&A', '210100 Accrued Bonus', 2400000, 'COMP-2026-11-018', ''),
    ('JE-2026-11-0040', date(2026, 11, 30), 'Monday', 'Maria Lopez', 'Performance marketing accrual — Meta/Google/TikTok', '600100 Marketing - Performance', '210400 Accrued Marketing', 8450000, 'MKT-2026-11-031', ''),
    ('JE-2026-11-0041', date(2026, 11, 30), 'Monday', 'David Park', 'CDN cost accrual — November traffic', '500200 Streaming Delivery (CDN)', '210500 Other Accruals', 800000, 'INFRA-2026-11-007', ''),
    ('JE-2026-11-0042', date(2026, 11, 28), 'Saturday', 'James Walker', 'Reclassification — content costs', '500100 Content Amortization', '600100 Marketing - Performance', 1500000, '', 'REVIEW'),  # FLAG
    ('JE-2026-11-0043', date(2026, 11, 30), 'Monday', 'Sarah Chen', 'D&O insurance amortization', '600700 Professional Fees', '120300 Prepaid Insurance', 200000, 'INS-2026-11-001', ''),
    ('JE-2026-11-0044', date(2026, 11, 30), 'Monday', 'Maria Lopez', 'Vacation accrual true-up', '600500 Salaries - G&A', '210200 Accrued Vacation', 400000, 'COMP-2026-11-019', ''),
    ('JE-2026-11-0045', date(2026, 11, 30), 'Monday', 'David Park', 'Sales commission accrual', '600300 Sales Commissions', '210500 Other Accruals', 200000, 'COMP-2026-11-020', ''),
    ('JE-2026-11-0046', date(2026, 11, 30), 'Monday', 'Sarah Chen', 'FX revaluation — IC balances', '700300 FX Gain/Loss', '130100 IC Receivable - LuminaEMEA', 450000, 'FX-2026-11-001', ''),
]
for r, row in enumerate(je_data, start=2):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = ARIAL; cell.border = BOX
        if c == 2:
            cell.number_format = DATE_FMT
        if c == 8:
            cell.number_format = CURRENCY
        if c in (5, 6, 7):
            cell.alignment = LEFT
    if row[9] == 'REVIEW':
        for c in range(1, 11):
            ws.cell(row=r, column=c).fill = FLAG_FILL

total_row = len(je_data) + 2
ws.cell(row=total_row, column=1, value='TOTAL').font = ARIAL_BOLD
ws.cell(row=total_row, column=1).fill = TOTAL_FILL
ws.cell(row=total_row, column=8, value=f'=SUM(H2:H{total_row-1})').font = ARIAL_BOLD
ws.cell(row=total_row, column=8).number_format = CURRENCY
ws.cell(row=total_row, column=8).fill = TOTAL_FILL
ws.cell(row=total_row, column=8).border = BOX

for i, w in enumerate([18, 12, 13, 16, 50, 36, 36, 16, 22, 10], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ============================================================
# Save
# ============================================================
out_path = '/home/claude/close-system/lumina_close_dataset.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
